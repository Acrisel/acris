#!/usr/bin/env python3 
# -*- encoding: utf-8 -*-
##############################################################################
#
#    Acrisel LTD
#    Copyright (C) 2008- Acrisel (acrisel.com) . All Rights Reserved
#
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see http://www.gnu.org/licenses/.
#
##############################################################################
'''
Created on Mar 23, 2017

@author: arnon
'''
from openpyxl import load_workbook
import collections
import os

def get_range(table_range=None, sheet_name=None):
    if sheet_name is not None and isinstance(table_range, collections.Mapping):
            table_range=table_range.get(sheet_name)
            
    if table_range is None or len(table_range) == 0:
        table_range=((1,1), (0,0))
    start_coor, end_coor = table_range
    if len(start_coor) == 0:
        start_coor=(1,1)
    start_row, start_col=start_coor
    if len(end_coor) > 0:
        end_coor=(0,0)
    end_row, end_col=end_coor
    return start_row, start_col, end_row, end_col

def get_sep_line(sym='-', sizes=[]):
    content=[sym*s for s in sizes]
    return '+' + '+'.join(content) + '+' 

def cell2str(cell):
    value=cell.value
    if value is None: value=''
    return str(value)

def mkline(row, col_size):
    line_items=[ "%s%s" %(cell2str(cell), " "*(size-len(cell2str(cell))))for cell, size in zip(row, col_size)]
    return '|' + '|'.join(line_items) + '|\n'

def get_dest_base(file, dest=None, one_file=True):
    if dest is None:
        result=file.replace('.xlsx', '')
    else:
        result=dest.replace('.rst', '')
    return result

def xlsx2rst(file, dest=None, sheets=[], one_file=True, table_range=None, header_row=1, ):
    ''' converts xlsx to rst format
    
    Args:
        file: path to xlsx file
        sheet: if empty, all sheets will be converted.  otherwise, only specified sheet-names will be produced.
            note: sheets names are case sensitive.
        table_range: start end coordinates in the form of ((row1, col1), (row2, col2))
            if (row1, col1) is empty, (1,1) is assumed
            if (row2, col2) is empty, last field filled
        header_row: number of row off range to take as header
        one_file: if True, will produce one file with all sheet tables consolidate separated by sheet-name title.
            if False, a file per sheet will be produced.
        dest: path to destination, defaults to xlsx.rst.  if dest is not None and one_file is False, dest would be extended with sheet name.
    '''
    
    if not file.endswith('.xlsx'): raise Exception("file must ends with .xlsx: %s" % file)
    if dest is not None and not dest.endswith('.rst'): raise Exception("dest must ends with .rst: %s" % dest)
    wb = load_workbook(file)
    wbsheets=wb.get_sheet_names()
    
    if sheets is not None and len(sheets) > 0:
        src_sheets=list(set(wbsheets) & set(sheets))
        if len(src_sheets) == 0:
            raise Exception("sheets not found with workbook: %s" % sheets)
    else:
        src_sheets=wbsheets
        
    if len(src_sheets) == 0:
        raise Exception("workbook does not have sheets")
    
    dest_base=get_dest_base(file, dest, one_file)
    
    if one_file:
        dest=dest_base+'.rst'
        if os.path.isfile(dest):
            os.remove(dest)
        
    for sheet_name in src_sheets:
        # sheet=wb.get_sheet_by_name(first_sheet)
        ws=wb[sheet_name]
        start_row, start_col, end_row, end_col= get_range(table_range, sheet_name)
        row_count=0
        header=list()
        body=list()
        col_size=list()
        
        if not one_file:
            dest=dest_base+'.%s.rst' % sheet_name
            if os.path.isfile(dest):
                os.remove(dest)
        
        with open(dest, 'a') as destfile:
            for row in ws.iter_rows():
                row_count += 1
                if row_count < start_row: continue
                if end_row > 0 and row_count > end_row: break
                cell_size=[len(str(cell.value)) for cell in row]
                if len(col_size) > 0:
                    col_size=[max(cols, cells) for cols, cells in zip(col_size, cell_size)]
                else:
                    col_size=cell_size  
                     
                eff_row=row_count-start_row+1
                
                if eff_row <= header_row:
                    header.append(row)
                else:
                    body.append(row)
                    
            body_sep=get_sep_line(sym='-', sizes=col_size)
            header_sep=get_sep_line(sym='=', sizes=col_size)
            
            destfile.write("%s\n%s\n\n" % (sheet_name, '='*len(sheet_name)) + '\n')
            destfile.write(body_sep+'\n')
            for row in header:
                destfile.write(mkline(row, col_size=col_size))
            destfile.write(header_sep+'\n')
            for row in body:
                destfile.write(mkline(row, col_size=col_size))
                destfile.write(body_sep+'\n')
                
if '__main__' == __name__:
    import argparse
    parser = argparse.ArgumentParser(description='Converts xlsx workbook into restructured text format')
    parser.add_argument("file", metavar='XLSX', 
                        help="xlsx files to convert")
    parser.add_argument("-o", "--output", metavar='RST', required=False, 
                        help="destination rst file")
    parser.add_argument("-s", "--sheet", metavar='SHEET', nargs='*', 
                        help="list of sheets; default to all available sheets")
    parser.add_argument("--start-row", metavar='NUMBER', type=int, nargs='?', default=1, dest='start_row',
                        help="table start row, defaults to 1")
    parser.add_argument("--end-row", metavar='NUMBER', type=int, nargs='?', default=1, dest='start_col',
                        help="table start col, defaults to 1")
    parser.add_argument("--start-col", metavar='NUMBER', type=int, nargs='?', default=0, dest='end_row',
                        help="table start row, defaults to 0")
    parser.add_argument("--end-col", metavar='NUMBER', type=int, nargs='?', default=0, dest='end_col',
                        help="table start col, defaults to 0")
    parser.add_argument("-r", "--header", metavar='NUMBER', type=int, nargs='?', default=1,
                        help="header row count")
    parser.add_argument("--one-file", action='store_true', default=False,
                        help="when set, single file is created")
    
    args = parser.parse_args()
    
    table_range=((args.start_row, args.start_col), (args.end_row, args.end_col))
    xlsx2rst(file=args.file, dest=args.output, sheets=args.sheet, table_range=table_range, header_row=args.header, one_file=args.one_file)