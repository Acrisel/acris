#!/bin/env python3
# -*- encoding: utf-8 -*-
##############################################################################
#
#    Acrisel LTD
#    Copyright (C) 2008- Acrisel (acrisel.com) . All Rights Reserved
#
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see http://www.gnu.org/licenses/.
#
##############################################################################

import argparse
import csv
import sys
import os
from openpyxl.workbook import Workbook

parser = argparse.ArgumentParser(description='Creates Excel file from one or more CSV files.  If multiple CSV are provided, they wiull be mapped to separated sheets.  If "-" is provided, input will be acquire from stdin.')
parser.add_argument("input_files", metavar='CSV', nargs='+', help="csv files to merge in xlsx; if -, stdin is assumed")
parser.add_argument("-d", "--delimiter", metavar='DELIMITER', default=",", help="select delimiter character")
parser.add_argument("-o", "--out", metavar='OUTFILE', help="output xlsx filename")
args = parser.parse_args()


wb = Workbook()
first_sheet=True


for input_file in args.input_files:
    if input_file == '-':
        input_fh=sys.stdin
        sheet_name='sheet'
    elif not input_file.endswith(".csv"):
        sys.stderr.write("Error: File does not have the ending \".csv\".\n")
        sys.exit(2)
    else:
        input_fh=open(input_file)
        sheet_name=os.path.basename(input_file).rpartition('.')[0]

    if first_sheet:
        sheet = wb.active
        sheet.title=sheet_name
        first_sheet=False
    else:
        sheet = wb.create_sheet(sheet_name)

    for row_index, row in enumerate(
        csv.reader(input_fh, delimiter=args.delimiter), start=1):
        for col_index, col in enumerate(row, start=1):
            sheet.cell(row = row_index, column = col_index).value = col

wb.save(args.out )
